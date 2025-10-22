#!/usr/bin/env python3
"""
qa_to_text_chunks.py

Convert a Q&A source (Excel .xlsx or chat-style JSONL) into plain text chunk files:

  data/chunks/<source_basename>/chunk_001.txt
  each file contains:
    Q: {question}

    A: {answer}

Usage:
  python qa_to_text_chunks.py --input "VNU_Q&A.xlsx"
  python qa_to_text_chunks.py --input train_chat.jsonl --out-root data/chunks

Options:
  --input / -i    : input file (.xlsx or .jsonl) (required)
  --out-root / -o : output root directory (default: data/chunks)
  --sheet / -s    : sheet name or index for Excel (optional)
  --qcol          : explicit question column name (optional)
  --acol          : explicit answer column name (optional)
"""
import argparse
import json
from pathlib import Path
import unicodedata
import pandas as pd
from openpyxl import load_workbook

# ---------- helpers ----------
def normalize_header(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ASCII", "ignore").decode("ASCII")
    return s.lower().strip()

QUESTION_KEYS = {"cau hoi", "cauhoi", "cau_hoi", "cau-hoi", "question", "q", "quest"}
ANSWER_KEYS   = {"tra loi", "traloi", "tra_loi", "tra-loi", "answer", "a", "ans", "tra-loi"}

def detect_q_a_columns(df: pd.DataFrame, qcol_hint=None, acol_hint=None):
    if qcol_hint and acol_hint and qcol_hint in df.columns and acol_hint in df.columns:
        return qcol_hint, acol_hint
    normalized = {col: normalize_header(col) for col in df.columns}
    qcol = None
    acol = None
    for col, norm in normalized.items():
        tokens = set(norm.replace("-", " ").replace("_", " ").split())
        if tokens & QUESTION_KEYS and qcol is None:
            qcol = col
        if tokens & ANSWER_KEYS and acol is None:
            acol = col
    if qcol is None or acol is None:
        # fallback to first two columns
        if len(df.columns) >= 2:
            qcol = qcol or df.columns[0]
            acol = acol or df.columns[1]
    return qcol, acol

def _worksheet_to_dataframe(ws):
    # Convert an openpyxl worksheet to a DataFrame using first non-empty row as header
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return pd.DataFrame()
    header_row = None
    header_idx = None
    for idx, r in enumerate(values):
        if any(cell is not None and str(cell).strip() != "" for cell in r):
            header_row = r
            header_idx = idx
            break
    if header_row is None:
        return pd.DataFrame()
    headers = [
        (str(h).strip() if h is not None and str(h).strip() != "" else f"col_{i+1}")
        for i, h in enumerate(header_row)
    ]
    data_rows = []
    for r in values[header_idx + 1 :]:
        # trim or pad row to header length
        row_vals = [None] * len(headers)
        for i, cell in enumerate(r):
            if i < len(headers):
                # Format numbers to match Excel display (round to 2 decimal places)
                if isinstance(cell, (int, float)):
                    if isinstance(cell, float) and cell.is_integer():
                        row_vals[i] = int(cell)
                    else:
                        row_vals[i] = round(cell, 2)
                else:
                    row_vals[i] = cell
        data_rows.append(row_vals)
    return pd.DataFrame(data_rows, columns=headers)


def read_excel_input(path: Path, sheet=None, qcol=None, acol=None):
    # Read with openpyxl data_only=True to get evaluated values instead of formulas
    wb = load_workbook(path, data_only=True)
    sheets = {}
    if sheet is not None:
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[int(sheet)]
        sheets[ws.title] = _worksheet_to_dataframe(ws)
    else:
        for ws in wb.worksheets:
            sheets[ws.title] = _worksheet_to_dataframe(ws)
    return sheets

def read_chat_jsonl_input(path: Path):
    rows = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            msgs = obj.get("messages") or []
            user = None
            assistant = None
            for m in msgs:
                if not isinstance(m, dict):
                    continue
                role = m.get("role")
                if role == "user" and not user:
                    user = m.get("content", "").strip()
                if role == "assistant" and not assistant:
                    # prefer assistant content (skip assistant entries that only have tool_calls and no content)
                    if m.get("content"):
                        assistant = m.get("content", "").strip()
            if user and assistant:
                rows.append({"question": user, "answer": assistant, "source": str(path)})
    df = pd.DataFrame(rows)
    return df, "question", "answer"

# ---------- main ----------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", "-i", required=True, help="Input file (.xlsx or .jsonl)")
    parser.add_argument("--out-root", "-o", default="data/chunks", help="Output root directory (default: data/chunks)")
    parser.add_argument("--sheet", "-s", default=None, help="Excel sheet name or index (optional)")
    parser.add_argument("--qcol", default=None, help="Question column name (optional)")
    parser.add_argument("--acol", default=None, help="Answer column name (optional)")
    args = parser.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        raise SystemExit(f"Input not found: {inp}")

    out_root = Path(args.out_root)
    source_basename = inp.stem  # filename without suffix
    target_dir = out_root / source_basename
    target_dir.mkdir(parents=True, exist_ok=True)

    # read input to DataFrame(s) of rows with columns question/answer
    if inp.suffix.lower() in [".xlsx", ".xls"]:
        sheet_to_df = read_excel_input(inp, args.sheet, args.qcol, args.acol)
    else:
        df, qcol, acol = read_chat_jsonl_input(inp)

    if inp.suffix.lower() in [".xlsx", ".xls"]:
        # For Excel: create separate chunk folders per sheet
        overall_total = 0
        for sheet_name, df in sheet_to_df.items():
            if df.empty:
                continue
            qcol, acol = detect_q_a_columns(df, args.qcol, args.acol)
            if qcol is None or acol is None:
                continue
            pairs = []
            for _, row in df.iterrows():
                q = (row.get(qcol) if qcol in row else row.get("question")) or ""
                a = (row.get(acol) if acol in row else row.get("answer")) or ""
                q = str(q).strip()
                a = str(a).strip()
                if not q or not a:
                    continue
                pairs.append((q, a))
            if not pairs:
                continue
            # sanitize sheet name for filesystem
            safe_sheet = "".join(ch if ch.isalnum() or ch in ("-", "_", " ") else "_" for ch in sheet_name).strip()
            sheet_dir = target_dir / safe_sheet
            sheet_dir.mkdir(parents=True, exist_ok=True)
            pad = max(3, len(str(len(pairs))))
            for i, (q, a) in enumerate(pairs, start=1):
                filename = f"chunk_{i:0{pad}d}.txt"
                path = sheet_dir / filename
                content = f"Q: {q}\n\nA: {a}"
                path.write_text(content, encoding="utf-8")
            print(f"Wrote {len(pairs)} chunk files to: {sheet_dir}")
            overall_total += len(pairs)
        if overall_total == 0:
            print("No valid Q/A pairs found across sheets.")
            return
    else:
        if qcol is None or acol is None:
            raise SystemExit("Could not detect question/answer columns automatically. Provide --qcol and --acol.")

        pairs = []
        for _, row in df.iterrows():
            q = (row.get(qcol) if qcol in row else row.get("question")) or ""
            a = (row.get(acol) if acol in row else row.get("answer")) or ""
            q = str(q).strip()
            a = str(a).strip()
            if not q or not a:
                continue
            pairs.append((q, a))
        total = len(pairs)
        if total == 0:
            print("No valid Q/A pairs found.")
            return
        pad = max(3, len(str(total)))
        for i, (q, a) in enumerate(pairs, start=1):
            filename = f"chunk_{i:0{pad}d}.txt"
            path = target_dir / filename
            content = f"Q: {q}\n\nA: {a}"
            path.write_text(content, encoding="utf-8")
        print(f"Wrote {total} chunk files to: {target_dir}")

if __name__ == "__main__":
    main()
